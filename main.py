import tkinter as tk
from tkinter import ttk
from tkinter import messagebox
from tkinter import filedialog
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font


class Application:
    def __init__(self, root):
        self.root = root
        self.root.title("Kalkulator Materiałów")

        # Label i ComboBox dla wyboru pliku Excel
        self.file_label = ttk.Label(root, text="Wybierz plik Excel:")
        self.file_label.grid(row=0, column=0, padx=5, pady=5, sticky="w")
        self.file_entry = ttk.Entry(root, state="readonly")
        self.file_entry.grid(row=0, column=1, padx=5, pady=5, sticky="w")
        self.browse_button = ttk.Button(root, text="Przeglądaj", command=self.browse_file)
        self.browse_button.grid(row=0, column=2, padx=5, pady=5, sticky="w")

        # Pole do wyszukiwania materiałów
        self.search_label = ttk.Label(root, text="Wyszukaj materiał:")
        self.search_label.grid(row=1, column=0, padx=5, pady=5, sticky="w")
        self.search_entry = ttk.Entry(root)
        self.search_entry.grid(row=1, column=1, padx=5, pady=5, sticky="w")
        self.search_entry.bind("<KeyRelease>", self.filter_materials)

        # Lista dostępnych materiałów
        self.materials_label = ttk.Label(root, text="Dostępne materiały:")
        self.materials_label.grid(row=2, column=0, padx=5, pady=5, sticky="w")
        self.material_listbox = tk.Listbox(root, selectmode="single", exportselection=0)
        self.material_listbox.grid(row=2, column=1, padx=5, pady=5, sticky="nsew", columnspan=2)

        # Label i Entry dla wprowadzenia długości
        self.length_label = ttk.Label(root, text="Długość:")
        self.length_label.grid(row=3, column=0, padx=5, pady=5, sticky="w")
        self.length_entry = ttk.Entry(root)
        self.length_entry.grid(row=3, column=1, padx=5, pady=5, sticky="w")

        # Label i Entry dla wprowadzenia ilości
        self.quantity_label = ttk.Label(root, text="Ilość:")
        self.quantity_label.grid(row=4, column=0, padx=5, pady=5, sticky="w")
        self.quantity_entry = ttk.Entry(root)
        self.quantity_entry.grid(row=4, column=1, padx=5, pady=5, sticky="w")

        # Label i Entry dla wprowadzenia ceny jednostkowej
        self.unit_price_label = ttk.Label(root, text="Cena jednostkowa:")
        self.unit_price_label.grid(row=5, column=0, padx=5, pady=5, sticky="w")
        self.unit_price_entry = ttk.Entry(root)
        self.unit_price_entry.grid(row=5, column=1, padx=5, pady=5, sticky="w")

        # Przycisk do dodawania pozycji
        self.add_button = ttk.Button(root, text="Dodaj", command=self.add_item)
        self.add_button.grid(row=6, column=0, columnspan=3, padx=5, pady=5)

        # Przycisk do generowania raportu
        self.generate_button = ttk.Button(root, text="Generuj", command=self.generate_report)
        self.generate_button.grid(row=7, column=0, columnspan=3, padx=5, pady=5)

        # Inicjalizacja listy materiałów
        self.material_list = []
        self.filtered_materials = []

        # Inicjalizacja słownika do przechowywania wag teoretycznych
        self.theoretical_weights = {}

        # Inicjalizacja listy dodanych pozycji
        self.added_items = []

        # Inicjalizacja pola tekstowego do wyświetlania dodanych pozycji
        self.added_items_text = tk.Text(root, height=10, width=50)
        self.added_items_text.grid(row=8, column=0, columnspan=3, padx=5, pady=5)

        # Inicjalizacja pliku Excel
        self.wb = None
        self.ws = None

    def browse_file(self):
        file_path = filedialog.askopenfilename(filetypes=[("Pliki Excel", "*.xlsx")])
        if file_path:
            self.file_entry.config(state="normal")
            self.file_entry.delete(0, tk.END)
            self.file_entry.insert(0, file_path)
            self.file_entry.config(state="readonly")
            self.load_materials(file_path)

    def load_materials(self, file_path):
        try:
            wb = load_workbook(filename=file_path, data_only=True)
            ws = wb.active

            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=3):
                material = row[0].value
                theoretical_weight = row[1].value
                if material and theoretical_weight:
                    self.material_list.append(material)
                    self.theoretical_weights[material] = theoretical_weight

            self.material_listbox.delete(0, tk.END)
            for material in self.material_list:
                self.material_listbox.insert(tk.END, material)

            self.filtered_materials = self.material_list.copy()
        except Exception as e:
            messagebox.showerror("Błąd", f"Wystąpił błąd podczas wczytywania danych z pliku: {str(e)}")

    def filter_materials(self, event=None):
        search_term = self.search_entry.get().lower()
        self.filtered_materials = [material for material in self.material_list if search_term in material.lower()]

        self.material_listbox.delete(0, tk.END)
        for material in self.filtered_materials:
            self.material_listbox.insert(tk.END, material)

    def add_item(self):
        selected_material_index = self.material_listbox.curselection()
        if not selected_material_index:
            messagebox.showerror("Błąd", "Proszę wybrać materiał z listy.")
            return

        material = self.filtered_materials[selected_material_index[0]]
        length = self.length_entry.get()
        quantity = self.quantity_entry.get()
        unit_price = self.unit_price_entry.get()

        try:
            length = float(length)
            quantity = int(quantity)
            unit_price = float(unit_price)
        except ValueError:
            messagebox.showerror("Błąd", "Proszę wprowadzić poprawne wartości liczbowe.")
            return

        if length <= 0 or quantity <= 0 or unit_price <= 0:
            messagebox.showerror("Błąd", "Długość, ilość i cena jednostkowa muszą być większe od zera.")
            return

        theoretical_weight = self.theoretical_weights.get(material)
        if theoretical_weight is None:
            messagebox.showerror("Błąd", "Waga teoretyczna nie została znaleziona dla wybranego materiału.")
            return

        total_weight = theoretical_weight * length * quantity
        value = unit_price * total_weight

        # Dodanie pozycji do listy dodanych pozycji
        self.added_items.append(
            [material, round(length, 2), quantity, round(length * quantity, 2), round(unit_price, 2),
             round(theoretical_weight, 2), round(total_weight, 2), round(value, 2)])
        self.update_added_items_text()

        messagebox.showinfo("Sukces", "Pozycja dodana pomyślnie.")

    def update_added_items_text(self):
        self.added_items_text.delete("1.0", tk.END)
        for item in self.added_items:
            self.added_items_text.insert(tk.END,
                                         f"Materiał: {item[0]}, Długość: {item[1]}, Ilość: {item[2]}, Długość całkowita: {item[3]}, Cena jednostkowa: {item[4]}, Waga teoretyczna: {item[5]}, Waga całkowita: {item[6]}, Wartość: {item[7]}\n")

    def generate_report(self):
        if not self.added_items:
            messagebox.showerror("Błąd", "Brak danych do wygenerowania raportu.")
            return

        try:
            self.wb = Workbook()
            self.ws = self.wb.active
            self.ws.append(["Materiał", "Długość", "Ilość", "Długość całkowita", "Cena jednostkowa", "Waga teoretyczna",
                            "Waga całkowita", "Wartość"])
            for item in self.added_items:
                self.ws.append(item)

            # Styl nagłówków
            header_font = Font(bold=True)
            for cell in self.ws["1:1"]:
                cell.font = header_font

            # Przykładowe formatowanie kolumn
            for column in range(1, len(self.added_items[0]) + 1):
                self.ws.column_dimensions[chr(64 + column)].width = 15

            filename = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Plik Excel", "*.xlsx")])
            if filename:
                self.wb.save(filename)
                messagebox.showinfo("Sukces", "Raport został wygenerowany pomyślnie.")
        except Exception as e:
            messagebox.showerror("Błąd", f"Wystąpił błąd podczas generowania raportu: {str(e)}")


# Inicjalizacja aplikacji
root = tk.Tk()
app = Application(root)
root.mainloop()
